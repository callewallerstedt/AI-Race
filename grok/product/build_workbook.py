#!/usr/bin/env python3
"""Build the Svensk Frilans Ekonomi Pack (.xlsx).

Not tax advice and not a bookkeeping system. A practical cash, pricing,
and invoice-log workbook for Swedish freelancers and egenanställda.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).with_name("Svensk-Frilans-Ekonomi-Pack.xlsx")

NAVY = "1B2A4A"
TEAL = "0F766E"
LIGHT = "F0FDFA"
AMBER = "FEF3C7"
WHITE = "FFFFFF"
GRAY = "64748B"
RED = "FEE2E2"
GREEN = "DCFCE7"

thin = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def style_header(cell, fill=NAVY):
    cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin


def style_input(cell):
    cell.fill = PatternFill("solid", fgColor=AMBER)
    cell.border = thin
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_calc(cell):
    cell.fill = PatternFill("solid", fgColor=LIGHT)
    cell.border = thin


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_title(ws, text, merge="A1:G1"):
    ws.merge_cells(merge)
    cell = ws["A1"]
    cell.value = text
    cell.font = Font(name="Calibri", bold=True, size=18, color=NAVY)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28


def build_readme(wb: Workbook):
    ws = wb.active
    ws.title = "00_README"
    add_title(ws, "Svensk Frilans Ekonomi Pack — README", "A1:B1")
    rows = [
        ("Version", "1.0 — AI Race product build"),
        ("Purpose", "Practical cash control for Swedish freelancers / egenanställda"),
        ("Not", "Not bookkeeping software. Not tax advice. Not a substitute for Bokio/Fortnox/accountant."),
        ("Legal note", "Swedish bookkeeping law requires immutable bookkeeping systems; Excel is for planning/logs only."),
        ("Sheets", "Setup · Fee_Catalog · Pricing · Income_Log · Expense_Log · Cashflow · Invoice_Log · Tax_Estimate · Dashboard"),
        ("How to use", "1) Fill Setup. 2) Log income/expenses. 3) Price jobs in Pricing. 4) Review Dashboard weekly."),
        ("Google Sheets", "File → Import → Upload this .xlsx → Replace spreadsheet."),
        ("Currency", "SEK throughout"),
        ("Support", "Product includes formulas only; update fee rates if platforms change."),
    ]
    ws["A3"] = "Field"
    ws["B3"] = "Value"
    style_header(ws["A3"])
    style_header(ws["B3"])
    for i, (k, v) in enumerate(rows, 4):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
        ws[f"A{i}"].font = Font(bold=True, color=NAVY)
        ws[f"A{i}"].border = thin
        ws[f"B{i}"].border = thin
        ws[f"B{i}"].alignment = Alignment(wrap_text=True)
    set_widths(ws, [18, 100])
    ws.row_dimensions[7].height = 45
    ws.row_dimensions[8].height = 45


def build_setup(wb: Workbook):
    ws = wb.create_sheet("01_Setup")
    add_title(ws, "Setup — your defaults", "A1:C1")
    ws["A3"] = "Setting"
    ws["B3"] = "Value"
    ws["C3"] = "Notes"
    for col in "ABC":
        style_header(ws[f"{col}3"])

    settings = [
        ("Display name", "Ditt namn", "Shown on invoice log exports"),
        ("Default platform", "Direktkund", "Direktkund / Frilans Finans / Fiverr / Gumroad / Annat"),
        ("Default hourly target SEK", 650, "What you want to earn per effective hour before platform fees"),
        ("Target net monthly SEK", 25000, "After rough tax estimate — planning only"),
        ("Illustrative income tax %", 0.30, "Rough personal planning rate — replace with your own estimate"),
        ("Employer fee / social approx %", 0.3142, "Relevant when comparing egenanställning gross invoice vs net"),
        ("Frilans Finans fee %", 0.06, "Update if your plan differs"),
        ("Fiverr fee %", 0.20, "Seller commission"),
        ("Gumroad fee %", 0.10, "Approx platform fee; processing may differ"),
        ("Payhip free fee %", 0.05, "Plus separate Stripe/PayPal processing outside this model"),
        ("VAT registered?", "Nej", "Ja/Nej — if Ja, use 25% on eligible sales in Pricing"),
        ("Default VAT %", 0.25, "Standard Swedish moms"),
        ("Starting cash SEK", 200, "Optional opening balance for Cashflow"),
        ("Race / project tag", "AI-Race", "Optional label"),
    ]
    for i, (k, v, note) in enumerate(settings, 4):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
        ws[f"C{i}"] = note
        ws[f"A{i}"].font = Font(bold=True)
        style_input(ws[f"B{i}"])
        ws[f"C{i}"].font = Font(color=GRAY, italic=True)
        ws[f"A{i}"].border = thin
        ws[f"C{i}"].border = thin
        if isinstance(v, float) and v < 1:
            ws[f"B{i}"].number_format = "0.00%"
        elif isinstance(v, (int, float)):
            ws[f"B{i}"].number_format = "#,##0.00"

    ws["A20"] = "Named ranges used by other sheets read these B cells directly."
    ws["A20"].font = Font(italic=True, color=GRAY)
    set_widths(ws, [32, 22, 70])

    dv_platform = DataValidation(
        type="list",
        formula1='"Direktkund,Frilans Finans,Fiverr,Gumroad,Payhip,Annat"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_platform)
    dv_platform.add(ws["B5"])
    dv_vat = DataValidation(type="list", formula1='"Ja,Nej"', allow_blank=True)
    ws.add_data_validation(dv_vat)
    dv_vat.add(ws["B14"])


def build_fee_catalog(wb: Workbook):
    ws = wb.create_sheet("02_Fee_Catalog")
    add_title(ws, "Fee catalog — edit rates when platforms change", "A1:E1")
    headers = ["Platform", "Fee %", "Settlement note", "Typical payout lag", "Source / checked"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(3, i, h)
        style_header(cell, TEAL)

    rows = [
        ("Direktkund / Swish", 0.0, "Direct payment", "Immediate–1 day", "Frilans Finans Swish docs"),
        ("Frilans Finans invoice", 0.06, "Gross invoice fee before tax/employer charges", "≈5 bank days salary", "frilansfinans.se 2026"),
        ("Fiverr", 0.20, "Seller commission on order value", "14 days after completion (new sellers)", "Fiverr Help 2026"),
        ("Gumroad", 0.10, "Platform fee; payout holds apply", "7-day hold + weekly payout; new-account review", "gumroad.com/help 2026"),
        ("Payhip Free", 0.05, "Plus Stripe/PayPal processing separately", "Processor-connected / often fast", "Payhip pricing pages 2026"),
        ("Etsy digital", 0.065, "Plus listing $0.20 + processing", "New seller holds possible", "Etsy fee guides 2026"),
        ("Upwork", 0.10, "Varies by contract history", "Platform withdrawal rules", "Verify before quoting"),
    ]
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.border = thin
            if c == 2:
                cell.number_format = "0.00%"
                style_input(cell)
            elif c == 1:
                cell.font = Font(bold=True)
    set_widths(ws, [24, 12, 48, 40, 28])


def build_pricing(wb: Workbook):
    ws = wb.create_sheet("03_Pricing")
    add_title(ws, "Job pricing calculator — quote with fees in mind", "A1:F1")
    ws["A3"] = "Inputs (yellow)"
    ws["A3"].font = Font(bold=True, color=TEAL)

    labels = [
        (4, "Job name", "Landing page — Basic"),
        (5, "Estimated hours", 6),
        (6, "Hourly target SEK", "='01_Setup'!B6"),
        (7, "Desired net to you SEK", "=B5*B6"),
        (8, "Platform", "Frilans Finans"),
        (9, "Platform fee %", "=IF(B8=\"Fiverr\",'01_Setup'!B11,IF(B8=\"Gumroad\",'01_Setup'!B12,IF(B8=\"Payhip\",'01_Setup'!B13,IF(B8=\"Frilans Finans\",'01_Setup'!B10,0))))"),
        (10, "VAT applicable?", "='01_Setup'!B14"),
        (11, "VAT %", "=IF(B10=\"Ja\",'01_Setup'!B15,0)"),
    ]
    ws["A3"] = "Field"
    ws["B3"] = "Value"
    style_header(ws["A3"])
    style_header(ws["B3"])
    for row, label, val in labels:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = val
        ws[f"A{row}"].border = thin
        if row in (4, 5, 8):
            style_input(ws[f"B{row}"])
        else:
            style_calc(ws[f"B{row}"])
        if row in (5,):
            ws[f"B{row}"].number_format = "0.00"
        if row in (6, 7):
            ws[f"B{row}"].number_format = "#,##0.00"
        if row in (9, 11):
            ws[f"B{row}"].number_format = "0.00%"

    ws["A13"] = "Outputs"
    ws["A13"].font = Font(bold=True, size=14, color=NAVY)
    outputs = [
        (14, "Gross to charge before VAT", "=IF((1-B9)=0,0,B7/(1-B9))"),
        (15, "Platform fee SEK", "=B14*B9"),
        (16, "VAT SEK", "=B14*B11"),
        (17, "Customer pays (incl VAT)", "=B14+B16"),
        (18, "You keep before income tax (approx)", "=B14-B15"),
        (19, "Illustrative after income tax", "=B18*(1-'01_Setup'!B8)"),
        (20, "Effective hourly after platform fee", "=IF(B5=0,0,B18/B5)"),
    ]
    ws["A13"] = "Output"
    ws["B13"] = "SEK / value"
    style_header(ws["A13"], TEAL)
    style_header(ws["B13"], TEAL)
    for row, label, formula in outputs:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = formula
        ws[f"A{row}"].border = thin
        style_calc(ws[f"B{row}"])
        ws[f"B{row}"].number_format = "#,##0.00"
        ws[f"B{row}"].font = Font(bold=True, color=NAVY)

    ws["D3"] = "Quick quote presets (edit freely)"
    ws["D3"].font = Font(bold=True, color=TEAL)
    presets = [
        ("Offer", "Hours", "List price SEK"),
        ("Webbfix Mini", 2, 499),
        ("Landing Basic", 6, 999),
        ("Landing Pro", 12, 1799),
        ("Sheet automation", 4, 799),
        ("Digital product", 0, 149),
    ]
    for i, row in enumerate(presets, 4):
        for c, val in enumerate(row, 4):
            cell = ws.cell(i, c, val)
            cell.border = thin
            if i == 4:
                style_header(cell, TEAL)
            elif c == 6:
                cell.number_format = "#,##0"
                style_input(cell)
            elif c == 5:
                style_input(cell)

    dv = DataValidation(
        type="list",
        formula1='"Direktkund,Frilans Finans,Fiverr,Gumroad,Payhip,Annat"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(ws["B8"])
    set_widths(ws, [36, 18, 4, 22, 10, 14])


def build_income_log(wb: Workbook):
    ws = wb.create_sheet("04_Income_Log")
    add_title(ws, "Income log — settled and pending", "A1:L1")
    headers = [
        "Date",
        "Client / source",
        "Description",
        "Platform",
        "Gross SEK",
        "Fee %",
        "Fee SEK",
        "Net SEK",
        "Status",
        "Settled date",
        "Evidence ref",
        "Notes",
    ]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(3, i, h), TEAL)

    # Sample starter row formulas for row 4..53
    for r in range(4, 54):
        ws[f"F{r}"] = (
            f'=IF(D{r}="","",IF(D{r}="Fiverr",\'01_Setup\'!B11,'
            f'IF(D{r}="Gumroad",\'01_Setup\'!B12,'
            f'IF(D{r}="Payhip",\'01_Setup\'!B13,'
            f'IF(D{r}="Frilans Finans",\'01_Setup\'!B10,0)))))'
        )
        ws[f"G{r}"] = f'=IF(E{r}="","",E{r}*F{r})'
        ws[f"H{r}"] = f'=IF(E{r}="","",E{r}-G{r})'
        for col in range(1, 13):
            ws.cell(r, col).border = thin
        for col in (5, 7, 8):
            ws.cell(r, col).number_format = "#,##0.00"
        ws[f"F{r}"].number_format = "0.00%"
        style_input(ws[f"A{r}"])
        style_input(ws[f"B{r}"])
        style_input(ws[f"C{r}"])
        style_input(ws[f"D{r}"])
        style_input(ws[f"E{r}"])
        style_calc(ws[f"F{r}"])
        style_calc(ws[f"G{r}"])
        style_calc(ws[f"H{r}"])
        style_input(ws[f"I{r}"])
        style_input(ws[f"J{r}"])
        style_input(ws[f"K{r}"])
        style_input(ws[f"L{r}"])

    dv_plat = DataValidation(
        type="list",
        formula1='"Direktkund,Frilans Finans,Fiverr,Gumroad,Payhip,Etsy,Annat"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_plat)
    dv_plat.add("D4:D53")
    dv_status = DataValidation(
        type="list",
        formula1='"Pending,Settled,Refunded,Lost"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_status)
    dv_status.add("I4:I53")

    ws["A55"] = "Totals"
    ws["E55"] = "=SUMIF(I4:I53,\"Settled\",E4:E53)"
    ws["G55"] = "=SUMIF(I4:I53,\"Settled\",G4:G53)"
    ws["H55"] = "=SUMIF(I4:I53,\"Settled\",H4:H53)"
    ws["A56"] = "Pending net"
    ws["H56"] = "=SUMIF(I4:I53,\"Pending\",H4:H53)"
    for cell in ("E55", "G55", "H55", "H56"):
        ws[cell].number_format = "#,##0.00"
        ws[cell].font = Font(bold=True)
    set_widths(ws, [12, 18, 28, 16, 12, 10, 10, 12, 12, 12, 16, 24])


def build_expense_log(wb: Workbook):
    ws = wb.create_sheet("05_Expense_Log")
    add_title(ws, "Expense log — every budget hit", "A1:H1")
    headers = ["Date", "Vendor", "Description", "Category", "Amount SEK", "Status", "Evidence", "Notes"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(3, i, h))
    for r in range(4, 54):
        for c in range(1, 9):
            cell = ws.cell(r, c)
            cell.border = thin
            if c in (1, 2, 3, 4, 5, 6, 7, 8):
                style_input(cell)
        ws[f"E{r}"].number_format = "#,##0.00"
    dv = DataValidation(
        type="list",
        formula1='"Software,Ads,Domain,Fees,Materials,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add("D4:D53")
    dv2 = DataValidation(type="list", formula1='"Planned,Paid,Refunded"', allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add("F4:F53")
    ws["A55"] = "Paid total"
    ws["E55"] = '=SUMIF(F4:F53,"Paid",E4:E53)'
    ws["E55"].number_format = "#,##0.00"
    ws["E55"].font = Font(bold=True)
    set_widths(ws, [12, 18, 32, 14, 12, 12, 16, 24])


def build_cashflow(wb: Workbook):
    ws = wb.create_sheet("06_Cashflow")
    add_title(ws, "30-day cashflow planner", "A1:F1")
    headers = ["Day", "Date", "Inflows SEK", "Outflows SEK", "Net SEK", "Running balance"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(3, i, h), TEAL)
    ws["A4"] = 0
    ws["B4"] = "Start"
    ws["C4"] = 0
    ws["D4"] = 0
    ws["E4"] = "=C4-D4"
    ws["F4"] = "='01_Setup'!B16+E4"
    for r in range(5, 35):
        ws[f"A{r}"] = r - 4
        ws[f"B{r}"] = f'=IF(B4="Start","",B4)'  # user fills dates
        style_input(ws[f"B{r}"])
        style_input(ws[f"C{r}"])
        style_input(ws[f"D{r}"])
        ws[f"E{r}"] = f"=C{r}-D{r}"
        ws[f"F{r}"] = f"=F{r-1}+E{r}"
        for c in range(1, 7):
            ws.cell(r, c).border = thin
        for c in (3, 4, 5, 6):
            ws.cell(r, c).number_format = "#,##0.00"
        style_calc(ws[f"E{r}"])
        style_calc(ws[f"F{r}"])
    for c in range(1, 7):
        ws.cell(4, c).border = thin
        if c >= 3:
            ws.cell(4, c).number_format = "#,##0.00"
    # Conditional formatting for negative balance
    ws.conditional_formatting.add(
        "F4:F34",
        FormulaRule(formula=["F4<0"], fill=PatternFill("solid", fgColor=RED)),
    )
    ws.conditional_formatting.add(
        "F4:F34",
        FormulaRule(formula=["F4>=0"], fill=PatternFill("solid", fgColor=GREEN)),
    )
    chart = LineChart()
    chart.title = "Running balance (SEK)"
    chart.style = 10
    chart.y_axis.title = "SEK"
    chart.x_axis.title = "Day"
    data = Reference(ws, min_col=6, min_row=3, max_row=34)
    cats = Reference(ws, min_col=1, min_row=4, max_row=34)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, "H3")
    set_widths(ws, [8, 14, 14, 14, 12, 16])


def build_invoice_log(wb: Workbook):
    ws = wb.create_sheet("07_Invoice_Log")
    add_title(ws, "Invoice / offer log", "A1:J1")
    headers = [
        "Invoice #",
        "Date",
        "Customer",
        "Description",
        "Amount ex VAT",
        "VAT",
        "Total",
        "Due date",
        "Paid?",
        "Channel",
    ]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(3, i, h))
    for r in range(4, 34):
        ws[f"A{r}"] = f'=IF(C{r}="","",CONCATENATE("INV-",TEXT({r-3},"000")))'
        ws[f"F{r}"] = f'=IF(E{r}="","",IF(\'01_Setup\'!B14="Ja",E{r}*\'01_Setup\'!B15,0))'
        ws[f"G{r}"] = f'=IF(E{r}="","",E{r}+F{r})'
        for c in range(1, 11):
            ws.cell(r, c).border = thin
        for c in (2, 3, 4, 5, 8, 9, 10):
            style_input(ws.cell(r, c))
        style_calc(ws[f"A{r}"])
        style_calc(ws[f"F{r}"])
        style_calc(ws[f"G{r}"])
        for c in (5, 6, 7):
            ws.cell(r, c).number_format = "#,##0.00"
    dv = DataValidation(type="list", formula1='"No,Yes,Overdue"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("I4:I33")
    set_widths(ws, [12, 12, 20, 28, 14, 10, 12, 12, 10, 16])


def build_tax_estimate(wb: Workbook):
    ws = wb.create_sheet("08_Tax_Estimate")
    add_title(ws, "Illustrative tax / take-home estimate — NOT advice", "A1:C1")
    ws["A3"] = "Warning"
    ws["B3"] = (
        "These figures are planning illustrations only. Swedish tax depends on municipality, "
        "deductions, whether you are employee via egenanställning, F-skatt, VAT, and more. "
        "Use an accountant or Skatteverket tools for real decisions."
    )
    ws["B3"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("B3:C3")
    ws.row_dimensions[3].height = 60

    ws["A5"] = "Settled gross income (from Income_Log)"
    ws["B5"] = "='04_Income_Log'!E55"
    ws["A6"] = "Settled platform fees"
    ws["B6"] = "='04_Income_Log'!G55"
    ws["A7"] = "Settled net before income tax"
    ws["B7"] = "='04_Income_Log'!H55"
    ws["A8"] = "Paid expenses"
    ws["B8"] = "='05_Expense_Log'!E55"
    ws["A9"] = "Illustrative taxable base"
    ws["B9"] = "=MAX(0,B7-B8)"
    ws["A10"] = "Illustrative income tax %"
    ws["B10"] = "='01_Setup'!B8"
    ws["A11"] = "Illustrative tax"
    ws["B11"] = "=B9*B10"
    ws["A12"] = "Illustrative take-home"
    ws["B12"] = "=B9-B11"
    for r in range(5, 13):
        ws[f"A{r}"].border = thin
        style_calc(ws[f"B{r}"])
        if r == 10:
            ws[f"B{r}"].number_format = "0.00%"
        else:
            ws[f"B{r}"].number_format = "#,##0.00"
    ws["B12"].font = Font(bold=True, size=14, color=TEAL)
    set_widths(ws, [40, 18, 40])


def build_dashboard(wb: Workbook):
    ws = wb.create_sheet("09_Dashboard")
    add_title(ws, "Dashboard — weekly review", "A1:B1")
    metrics = [
        (3, "Starting / tracked cash (Setup)", "='01_Setup'!B16"),
        (4, "Settled revenue net of platform fees", "='04_Income_Log'!H55"),
        (5, "Pending revenue net", "='04_Income_Log'!H56"),
        (6, "Paid expenses", "='05_Expense_Log'!E55"),
        (7, "Implied cash if pending ignores holds", "=B3+B4-B6"),
        (8, "Illustrative take-home", "='08_Tax_Estimate'!B12"),
        (9, "Open invoices total", "=SUMIF('07_Invoice_Log'!I4:I33,\"No\",'07_Invoice_Log'!G4:G33)"),
    ]
    ws["A2"] = "Metric"
    ws["B2"] = "SEK"
    style_header(ws["A2"])
    style_header(ws["B2"])
    for r, label, formula in metrics:
        ws[f"A{r}"] = label
        ws[f"B{r}"] = formula
        ws[f"A{r}"].border = thin
        style_calc(ws[f"B{r}"])
        ws[f"B{r}"].number_format = "#,##0.00"
    ws["A11"] = "Weekly checklist"
    ws["A11"].font = Font(bold=True, size=13, color=NAVY)
    checks = [
        "Update Income_Log statuses (Pending → Settled only when withdrawable)",
        "Log every expense the day it hits",
        "Re-price next offers using Pricing sheet fee logic",
        "Chase unpaid invoices > 7 days",
        "Do not count pending marketplace clears as race score",
    ]
    for i, t in enumerate(checks, 12):
        ws[f"A{i}"] = f"☐ {t}"
    set_widths(ws, [70, 16])


def main():
    wb = Workbook()
    build_readme(wb)
    build_setup(wb)
    build_fee_catalog(wb)
    build_pricing(wb)
    build_income_log(wb)
    build_expense_log(wb)
    build_cashflow(wb)
    build_invoice_log(wb)
    build_tax_estimate(wb)
    build_dashboard(wb)
    wb.save(OUT)
    print(f"Wrote {OUT} ({OUT.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
